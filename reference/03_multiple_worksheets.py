import os
import platform

def clear_screen():
    """
    Clears the terminal screen to make it easier to follow along with code.
    """
    if platform.system() == 'Windows':
        os.system('cls')
    else:
        os.system('clear')

clear_screen()

# ===================
# MULTIPLE WORKSHEETS
# ===================

'''
OVERVIEW
--------
Each Workbook can have multiple Worksheets (also just called Sheets).

'''
import openpyxl
my_workbook = openpyxl.Workbook()

# 1. FIND TITLE OF ACTIVE WORKSHEET
# You can use .title to get the title of a sheet.
# What is the title of the active sheet?
print(my_workbook.active.title)

# 2. CHANGE THE TITLE OF A WORKSHEET
# Change the title of the default sheet to "Students"
my_workbook.active.title = "Students"

# 3. ADD A NEW SHEET
# Create another sheet called "Classes" using .create_sheet("Sheet name")
my_workbook.create_sheet("Classes")

'''
TIP:
----
You can also remove a sheet using .remove, but we won't do that right now.
'''

# 4. CHANGE THE ACTIVE SHEET
# Change the active sheet to the "Classes" sheet. You can access a specific
# sheet using my_workbook["Sheet Name"]
my_workbook.active = my_workbook["Classes"]

# 5. ADD CONTENT TO THE ACTIVE SHEET
# Add some data to the first row or 2 of the active sheet.
my_workbook.active["A1"] = "Class Name"
my_workbook.active["B1"] = "Department"
my_workbook.active["A2"] = "IS 303"
my_workbook.active["B2"] = "Information Systems"

# 6. ADD CONTENT TO A SHEET BY REFERENCING THE SHEET'S NAME\
# Add some data to the "Students" sheet without making it the active sheet
my_workbook["Students"]["A1"] = "Name"
my_workbook["Students"]["B1"] = "Grade"
my_workbook["Students"]["A2"] = "Jimmy"
my_workbook["Students"]["B2"] = 3.8

# 7. STORE A CELL IN A VARIABLE
# Store "A2" from the Students sheet to a variable.
cell = my_workbook["Students"]["A2"]

# 8. ACCESS CELL ATTRIBUTES
# Print out the .value, .row, .column, and .coordinate of the cell
print(cell.value)
print(cell.row)
print(cell.column) # Notice that it gives a number instead of a letter
print(cell.coordinate)

# 9. ACCESS CELLS THROUGH .cell() METHOD
# Print out the value of the 1st column, 2nd row of the active sheet using the
# cells() method.
print(my_workbook.active.cell(1,2).value) 

# 10. ADD NEW VALUE THROUGH cell() METHOD
# Use cell() to add a new value to the active sheet at the third row,
# first column.
my_workbook.active.cell(3, 1, "New value")

# 11. FIND THE FURTHEST ROW AND FURTHEST COLUMN WITH DATA
# Use .max_row to get the furthest row with data, same with .max_col and print
# them out.
print(my_workbook.active.max_row)
print(my_workbook.active.max_column)

# 12. SAVE THE WORKBOOK
# Save the workbook as example_03.xlsx

my_workbook.save(filename="example_03.xlsx") 
my_workbook.close()