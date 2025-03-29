from helper_functions import clear_screen
clear_screen() # This just clears the terminal window each time we run code

# ===============================
# LOOPING THROUGH SHEETS PRACTICE
# ===============================

# 1. PRACTICE LOOPING
'''
Import the mock_grades file.

PART 1:
    Change the grade of anyone with a "C-" to be an "F" on the active sheet.
    Save the results as a new workbook called "example_06.xlsx"

PART 2:
    If you can successfully do that, then alter your code to change "C-" to "F"
    on ALL of the sheets in mock_grades. Save the results to
    "example_06.xlsx"

'''

import openpyxl

# load in mock_grades again
external_workbook = openpyxl.load_workbook("mock_grades.xlsx")

# PART 1:
# optionally store the active worksheet in a new variable
current_sheet = external_workbook.active

for row in current_sheet.iter_rows():
    for cell in row:
        if cell.value == "C-":
            cell.value = "F"

external_workbook.save("example_06.xlsx")      


# PART 2:
for worksheet in external_workbook.worksheets:
    for row in worksheet.iter_rows():
        if row[1].value == "C-":
            row[1].value = "F"

external_workbook.save("example_06.xlsx")      
external_workbook.close()