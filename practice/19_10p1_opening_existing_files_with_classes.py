# optional stuff that will clear the window each time you run it.
import os
import platform

def clear_screen():
    if platform.system() == 'Windows':
        os.system('cls')
    else:
        os.system('clear')

clear_screen()

###########################
# START READING HERE
###########################

'''
Write a program that can do the following:
    1. import books from excel sheet into book objects, all stored in a list
        Book objects should have instance variables that correspond to the excel sheet

    2. Allow the user to input additional data about another book, and then store it in the excel workbook
        (Note, you can store it in a new excel workbook just so you aren't overwriting the original in an incorrect way
        by accident)

'''
import openpyxl

class Book:
    def __init__(self, book_id, title, author, borrowed_by_id):
        self.book_id = book_id
        self.title = title
        self.author = author
        self.borrowed_by_id = borrowed_by_id

class Borrower:
    def __init__(self, borrower_id, name, email):
        self.borrower_id = borrower_id
        self.name = name
        self.email = email
        self.books_borrowed = []

    def borrow_book(self, book):
        if len(self.books_borrowed) < 3:
            self.books_borrowed.append(book)
        else:
            print("Can't borrow more than 3 books!")

# import all the Borrowers
library_workbook = openpyxl.load_workbook(r"CP14 - Python Working with Excel/library_system.xlsx")
book_list = []
for row in library_workbook["books"].iter_rows(values_only = True, min_row = 2):
    book = Book(*row) # this is using args. can also unpack the variables first, or just access them individually
    book_list.append(book)

borrower_list = []
for row in library_workbook["borrowers"].iter_rows(values_only = True, min_row = 2):
    borrower = Borrower(*row)
    borrower_list.append(borrower)

# note, if we were using an actual database, you could get around this when importing into python
# just by writing a better query to get the data in the first place.
for borrower in borrower_list:
    for book in book_list:
        if book.borrowed_by_id == borrower.borrower_id:
            borrower.borrow_book(book)


# now print out a section that will help you.
print("These are the books available to you: ")
for book in book_list:
    if book.borrowed_by_id is None:
        print(book.title)

chosen_book_title = input("enter a book title to check it out: ")

for book in book_list:
    if book.borrowed_by_id is None and chosen_book_title.lower().strip() == book.title.lower().strip():
        pass

# I'll finish this later


















