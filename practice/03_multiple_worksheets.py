from helper_functions import clear_screen
clear_screen() # This just clears the terminal window each time we run code

# ===================
# MULTIPLE WORKSHEETS
# ===================

'''
OVERVIEW
--------
Each Workbook can have multiple Worksheets (also just called Sheets).

'''
import openpyxl
my_workbook = openpyxl.Workbook()

# 1. FIND TITLE OF ACTIVE WORKSHEET
# You can use .title to get the title of a sheet.
# What is the title of the active sheet?


# 2. CHANGE THE TITLE OF A WORKSHEET
# Change the title of the default sheet to "Students"


# 3. ADD A NEW SHEET
# Create another sheet called "Classes" using .create_sheet("Sheet name")


'''
TIP:
----
You can also remove a sheet using .remove, but we won't do that right now.
'''

# 4. CHANGE THE ACTIVE SHEET
# Change the active sheet to the "Classes" sheet. You can access a specific
# sheet using my_workbook["Sheet Name"]


# 5. ADD CONTENT TO THE ACTIVE SHEET
# Add some data to the first row or 2 of the active sheet.


# 6. ADD CONTENT TO A SHEET BY REFERENCING THE SHEET'S NAME\
# Add some data to the "Students" sheet without making it the active sheet


# 7. STORE A CELL IN A VARIABLE
# Store "A2" from the Students sheet to a variable.


# 8. ACCESS CELL ATTRIBUTES
# Print out the .value, .row, .column, and .coordinate of the cell


# 9. ACCESS CELLS THROUGH .cell() METHOD
# Print out the value of the 1st column, 2nd row of the active sheet using the
# cells() method.


# 10. ADD NEW VALUE THROUGH cell() METHOD
# Use cell() to add a new value to the active sheet at the third row,
# first column.


# 11. FIND THE FURTHEST ROW AND FURTHEST COLUMN WITH DATA
# Use .max_row to get the furthest row with data, same with .max_col and print
# them out.


# 12. SAVE THE WORKBOOK
# Save the workbook as example_03.xlsx

