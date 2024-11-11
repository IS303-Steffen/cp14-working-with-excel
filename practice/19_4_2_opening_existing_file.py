# optional stuff that will clear the window each time you run it.
import os
import platform

def clear_screen():
    if platform.system() == 'Windows':
        os.system('cls')
    else:
        os.system('clear')

clear_screen()

###########################
# START READING HERE
###########################

# importing openpyxl import Workbook
import openpyxl

# you can open up an existing workbook by using openpyxl.load_workbook
# in the load_workbook function, you need to provide a file path to the excel spreadsheet
# I've given you an excel workbook called mock_grades. You can right click the file in the explorer
# window to the left to get the file path. Relative file paths are nice for it working on anyone's computer, but
# it depends on where you are running this python file from. Absolute paths will only work on your computer, but
# it doesn't matter where you are running the python file from.


# use .sheetnames to see the names of the sheets of the whole workbook


# that might come in handy when trying to do something to multiple tables.
# loop through each of the sheet names and print out the intidivudal sheet name


# if you want the entire worksheet object, not just the name of the worksheet,
# you can use .worksheets


# try printing out the value of A2 from each worksheet:


# .append
# use it to add a new row to the worksheet on the next empty row.
# try appending a list that has "New test student" and "A++"' as data


# save the file as mock_grades_altered and close the Workbook you opened.
