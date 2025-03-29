from helper_functions import clear_screen
clear_screen() # This just clears the terminal window each time we run code

# ===================
# LOAD EXISTING FILES
# ===================

'''
OVERVIEW
--------
Use the load_workbook() function to load in an already existing Excel Workbook
into Python.

RELATIVE FILE PATHS VS ABSOLUTE FILE PATHS
------------------------------------------
When you load in a file, you need to provide a file path. File paths can be
relative, meaning you only need to specify the file name and any folders
relative to the "working directory" meaning the folder that python is running
from. 

Absolute means you include the entire file path, which usually includes stuff
like your own computer username, etc. When possible use relative file paths.

'''

import openpyxl

# 1. LOAD IN mock_grades.xlsx
# Use .load_workbook to load in "mock_grades.xlsx", which is included in this
# repository. If you've only cloned this repository and don't have it opened
# with other files, you should be able to just put in the file name.
# Store the loaded Worksheet into a new variable.


# 2. PRINT OUT ALL THE SHEET NAMES
# Print out .sheetnames to see the names of the sheets of the whole workbook


# 3. LOOP THROUGH EACH SHEET NAME
# Use a for loop and print out each sheet name one at a time.


# 4. GET ENTIRE WORKSHEET OBJECTS
# Print out .worksheets. Then loop through each and print out the value of A2



# 5. ALTER EXISTING WORKSHEET WITH .append()
# Append list_to_append to the active sheet. .append() will just add data
# to the end of the worksheet.
list_to_append = ["New test student", "A++"]


# 6. SAVE A NEW COPY OF THE WORKBOOK
# Save the Workbook as example_04 and close the Workbook you opened.
