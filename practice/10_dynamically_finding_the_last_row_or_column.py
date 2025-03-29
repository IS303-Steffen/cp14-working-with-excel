from helper_functions import clear_screen
clear_screen() # This just clears the terminal window each time we run code

# ==========================================
# DYNAMICALLY FINDING THE LAST ROW OR COLUMN
# ==========================================

'''
OVERVIEW
--------
Sometimes you might want to add a filter, format something, etc but you don't
want to hardcode in a specific range. You can use
'''

import openpyxl
external_wb = openpyxl.load_workbook("library_system.xlsx")

# 1. FINDING THE MAX ROW
# Let's say you want to add a filter to colums A:D, but you don't know how many
# rows there are. You can use the .max_row property of any sheet to find the
# row number of the furthest row that has data on it. Find the max row of the
# "books" sheet, then apply a filter to that sheet without hardcoding the
# max row number. Save the workbook as example_10.xlsx

'''
Note about #2 and #3
----
On your project, you'll know exactly how many columns are in the workbooks
you're working with, so you don't need to use .max_column, but it is still a
generally useful thing to know.
'''

# 2. FINDING THE MAX COLUMN
# What if you don't know how many columns are in a sheet? You can use the 
# .max_column property of any sheet to find the column number of the furthest
# column that has data on it. Find the max column in the "borrowers" sheet
# and print it out.



# 3. TRANSLATING THE MAX COLUMN INTO A LETTER
# Notice how the max column is provided as a number? To use it as a cell
# reference (like A5, etc.) We need a the column represented as a letter
# instead of a number. Use the get_column_letter function from openpyxl.utils
# Then apply a filter to sheet "borrowers" starting in A1 all the way to 
# whatever the max column and max row is. Save the results as example_10.xlsx

from openpyxl.utils import get_column_letter



