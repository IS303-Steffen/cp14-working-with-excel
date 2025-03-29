from helper_functions import clear_screen
clear_screen() # This just clears the terminal window each time we run code

# ======================
# LOOPING THROUGH SHEETS
# ======================

'''
OVERVIEW
--------
Think of Worksheets as a list of rows, that each have a list of cells. When you
loop through a Worksheet, you need to get the row you want, then you can access
the cells.
'''

import openpyxl
# Loading in mock_grades
external_workbook = openpyxl.load_workbook("mock_grades.xlsx")
# optionally store the active worksheet in a new variable
current_sheet = external_workbook.active

# 1. USE .iter_rows() TO PRINT OUT EACH ROW
# .iter_rows() gives you each row. Use a for loop and print out each row in the
# active sheet. This isn't very useful yet.


# 2. USE THE values_only PARAMETER TO PRINT VALUES
# Use values_only = True to only return the values of the row.
# Print out the whole row


# 3. PRINT OUT EACH INDIVIDUAL VALUE
# Do the same as above, but use an inner loop to print out each value
# individually


# 4. GET CELL OBJECT TO ACCESS VALUE AND MORE
# Do the same thing, but now DO NOT use values_only = True.
# For every cell, print out the coordinate (use .coordinate on the cell)
# and the value (use .value)

'''
WHEN SHOULD YOU USE values_only=True?
-------------------------------------
If you just want to access the values, value_only=True is very convenient. If,
however, you want to change the values, or get access to more data about the
cell, you should not use values_only=True.
'''

# 5. SPECIFY A SUBSECTION OF THE WORKSHEET FOR .iter_rows()
# You can use min_row, min_col, max_col, max_row to only go through a section
# of the worksheet. Skip the header, and then just grab the next 5 rows of
# data in just the 2nd column


# 6. ANOTHER WAY TO GET A SUBSECTION:
# You can also get the specific cells you want by specifying a coordinate range
# Loop through each cell of ["B2:B6"] and print out the value

